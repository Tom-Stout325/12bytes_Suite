from __future__ import annotations

import json
import re
from pathlib import Path

from django.core.management.base import BaseCommand, CommandError


def _slugify(s: str) -> str:
    s = (s or "").strip().lower()
    s = re.sub(r"[^\w\s-]", "", s)
    s = re.sub(r"[\s_-]+", "-", s)
    s = re.sub(r"^-+|-+$", "", s)
    return s


def _is_req(v) -> bool:
    return str(v).strip().upper() == "X"


def _infer_account_type(name: str) -> str:
    n = (name or "").lower()

    if n in {"sales", "drone services", "photography services", "other income"}:
        return "income"
    if "returns" in n and "allowances" in n:
        return "income"
    if "sales tax collected" in n:
        return "liability"
    if n.startswith("equipment:"):
        return "asset"
    if n in {"depreciation", "depletion", "section 179", "amortization"}:
        return "journal"
    if n == "vehicle: loan payments":
        return "liability"
    if "interest" in n:
        return "expense"

    return "expense"


def _infer_contact_role(name: str, account_type: str, requires_contact: bool) -> str:
    if not requires_contact:
        return "any"
    n = (name or "").lower()
    if account_type == "income":
        return "customer"
    if "contract" in n:
        return "contractor"
    return "vendor"


class Command(BaseCommand):
    # help = "Build ledger/data/subcategory_rules.json from an Excel rules sheet (subcategory_list.xlsx)."

    def add_arguments(self, parser):
        parser.add_argument("--xlsx", type=str, required=True, help="Path to subcategory_list.xlsx")
        parser.add_argument(
            "--out",
            type=str,
            default=str(Path("ledger") / "data" / "subcategory_rules.json"),
            help="Output JSON path (default: ledger/data/subcategory_rules.json)",
        )

    def handle(self, *args, **opts):
        xlsx = Path(opts["xlsx"]).expanduser().resolve()
        out = Path(opts["out"]).expanduser().resolve()

        if not xlsx.exists():
            raise CommandError(f"Excel file not found: {xlsx}")

        try:
            import openpyxl
        except Exception as e:
            raise CommandError("openpyxl is required to read .xlsx files. Install it and retry.") from e

        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active

        headers = [ws.cell(1, c).value for c in range(1, 18)]
        if not headers or headers[0] != "Sub-Category":
            raise CommandError("Unexpected header row. Expected first column 'Sub-Category'.")

        rules: dict[str, dict] = {}

        for r in range(2, ws.max_row + 1):
            name = ws.cell(r, 1).value
            if not name:
                continue

            row = {headers[c - 1]: ws.cell(r, c).value for c in range(1, 18)}

            requires_contact = _is_req(row.get("contact"))
            requires_team = _is_req(row.get("team"))
            requires_job = _is_req(row.get("job"))
            requires_invoice_number = _is_req(row.get("invoice_number"))
            requires_receipt = _is_req(row.get("receipt"))
            requires_transport = _is_req(row.get("transport_type"))
            requires_vehicle = _is_req(row.get("vehicle"))
            requires_asset = _is_req(row.get("asset"))

            account_type = _infer_account_type(str(name))

            # Accounting overrides
            if name in {"Depreciation", "Depletion"}:
                requires_asset = True
                account_type = "journal"

            is_capitalizable = False
            auto_create_asset = False
            if str(name).startswith("Equipment:"):
                account_type = "asset"
                is_capitalizable = True
                auto_create_asset = True

            if name == "Contractors":
                requires_contact = True

            if name == "Sales Tax Collected":
                account_type = "liability"
                requires_receipt = False
                requires_team = False
                requires_job = False

            if name == "Vehicle: Loan Payments":
                account_type = "liability"
                requires_asset = True

            if name == "Vehicle: Loan Interest":
                account_type = "expense"

            if name in {"Sales", "Drone Services", "Returns & Allowances"}:
                account_type = "income"

            contact_role = _infer_contact_role(str(name), account_type, requires_contact)

            slug = _slugify(str(name))
            rule = {
                "name": str(name),
                "slug": slug,
                "account_type": account_type,
                "requires_contact": requires_contact,
                "contact_role": contact_role,
                "requires_team": requires_team,
                "requires_job": requires_job,
                "requires_invoice_number": requires_invoice_number,
                "requires_receipt": requires_receipt,
                "requires_transport": requires_transport,
                "requires_vehicle": requires_vehicle,
                "requires_asset": requires_asset,
                "is_capitalizable": is_capitalizable,
            }

            if str(name) == "Returns & Allowances":
                rule["auto_is_refund"] = True
            if auto_create_asset:
                rule["auto_create_asset"] = True

            rules[slug] = rule

        out.parent.mkdir(parents=True, exist_ok=True)
        payload = {
            "version": 1,
            "source": str(xlsx.name),
            "keying": "subcategory_slug (slugified name); applied by name or slug suffix match",
            "rules": rules,
        }
        out.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
        self.stdout.write(self.style.SUCCESS(f"Wrote {len(rules)} rules to {out}"))
