from __future__ import annotations

import csv
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO, StringIO
from typing import Any

from django.apps import apps
from django.db.models import Model, QuerySet
from django.http import HttpResponse
from django.utils import timezone
from django.utils.text import slugify


PROJECT_APP_LABELS = {
    "accounts",
    "assets",
    "contractor",
    "core",
    "invoices",
    "ledger",
    "vehicles",
}

MANUAL_BUSINESS_SCOPED_MODELS = {
    ("core", "Business"),
    ("core", "BusinessMembership"),
    ("accounts", "CompanyProfile"),
}

EXCLUDED_MODELS = {
    ("accounts", "Invitation"),  # Global invite records are not owned by one business.
}


@dataclass(frozen=True)
class BackupTableSpec:
    app_label: str
    model_name: str
    verbose_name: str
    table_name: str
    slug: str
    model: type[Model]

    @property
    def label(self) -> str:
        return f"{self.app_label}.{self.model_name}"


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, datetime):
        if timezone.is_aware(value):
            value = timezone.localtime(value)
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, Decimal):
        return f"{value:.2f}"
    return str(value)


def _safe_filename_part(value: str) -> str:
    return slugify(value or "moneypro") or "moneypro"


def _has_field(model: type[Model], field_name: str) -> bool:
    return any(getattr(field, "name", None) == field_name for field in model._meta.get_fields())


def business_backup_table_specs() -> list[BackupTableSpec]:
    """Return the MoneyPro tables that can be safely exported for one business.

    Most project data is tenant-owned with a `business` field. A few related records
    such as CompanyProfile and BusinessMembership are also business-scoped and are
    included explicitly. Global app/system tables are intentionally excluded so one
    tenant export does not leak another tenant's data.
    """

    specs: list[BackupTableSpec] = []

    for model in apps.get_models():
        app_label = model._meta.app_label
        model_name = model.__name__
        key = (app_label, model_name)

        if app_label not in PROJECT_APP_LABELS:
            continue
        if key in EXCLUDED_MODELS:
            continue

        include = _has_field(model, "business") or key in MANUAL_BUSINESS_SCOPED_MODELS
        if not include:
            continue

        specs.append(
            BackupTableSpec(
                app_label=app_label,
                model_name=model_name,
                verbose_name=str(model._meta.verbose_name_plural).title(),
                table_name=model._meta.db_table,
                slug=f"{app_label}-{slugify(model_name)}",
                model=model,
            )
        )

    return sorted(specs, key=lambda spec: (spec.app_label, spec.model_name))


def get_backup_table_spec_or_404(slug: str) -> BackupTableSpec:
    from django.shortcuts import get_object_or_404  # imported lazily only for Http404 behavior
    from django.http import Http404

    for spec in business_backup_table_specs():
        if spec.slug == slug:
            return spec
    raise Http404("Backup table not found.")


def queryset_for_business(spec: BackupTableSpec, business) -> QuerySet:
    model = spec.model

    if spec.app_label == "core" and spec.model_name == "Business":
        return model._default_manager.filter(pk=business.pk)

    if _has_field(model, "business"):
        return model._default_manager.filter(business=business)

    return model._default_manager.none()


def default_ordering(model: type[Model]) -> list[str]:
    field_names = {field.name for field in model._meta.concrete_fields}
    for candidate in ("date", "issue_date", "submitted_at", "generated_at", "created_at", "updated_at"):
        if candidate in field_names:
            return [f"-{candidate}", "-id"] if "id" in field_names else [f"-{candidate}"]
    return ["id"] if "id" in field_names else []


def table_columns(model: type[Model]) -> list[tuple[str, Any]]:
    """Return (header, getter) pairs for concrete model fields.

    Foreign keys include both the raw ID and a readable string value. File/Image
    fields export their storage key/path, not the file contents.
    """

    columns: list[tuple[str, Any]] = []

    for field in model._meta.concrete_fields:
        name = field.name
        internal_type = getattr(field, "get_internal_type", lambda: "")()

        if getattr(field, "is_relation", False) and getattr(field, "many_to_one", False):
            columns.append((f"{name}_id", lambda obj, n=name: getattr(obj, f"{n}_id", "")))
            columns.append((f"{name}_str", lambda obj, n=name: str(getattr(obj, n)) if getattr(obj, n, None) else ""))
            continue

        if internal_type in {"FileField", "ImageField"}:
            columns.append((name, lambda obj, n=name: getattr(getattr(obj, n, None), "name", "") or ""))
            continue

        columns.append((name, lambda obj, n=name: getattr(obj, n, "")))

    return columns


def rows_for_queryset(queryset: QuerySet) -> tuple[list[str], list[list[str]]]:
    columns = table_columns(queryset.model)
    headers = [header for header, _getter in columns]
    rows: list[list[str]] = []

    qs = queryset
    ordering = default_ordering(queryset.model)
    if ordering:
        qs = qs.order_by(*ordering)

    for obj in qs.iterator(chunk_size=1000):
        row: list[str] = []
        for _header, getter in columns:
            try:
                row.append(_format_value(getter(obj)))
            except Exception as exc:  # pragma: no cover - defensive export fallback
                row.append(f"[ERROR: {exc}]")
        rows.append(row)

    return headers, rows


def csv_response_for_table(*, spec: BackupTableSpec, business) -> HttpResponse:
    queryset = queryset_for_business(spec, business)
    headers, rows = rows_for_queryset(queryset)

    now = timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M%S")
    business_slug = _safe_filename_part(getattr(business, "slug", "") or getattr(business, "name", "business"))
    filename = f"{business_slug}-{spec.slug}-{now}.csv"

    response = HttpResponse(content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="{filename}"'

    writer = csv.writer(response)
    writer.writerow(headers)
    writer.writerows(rows)
    return response


def _unique_sheet_name(base_name: str, used: set[str]) -> str:
    safe = "".join(ch for ch in base_name if ch not in r"[]:*?/\\")[:31] or "Sheet"
    name = safe
    counter = 2
    while name in used:
        suffix = f" {counter}"
        name = f"{safe[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(name)
    return name


def workbook_bytes_for_business(*, business) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError as exc:  # pragma: no cover - only hit if dependency missing in deployment
        raise RuntimeError("openpyxl is required for Excel backup exports. Add openpyxl to requirements.txt and redeploy.") from exc

    specs = business_backup_table_specs()
    now_dt = timezone.localtime(timezone.now())
    now = now_dt.strftime("%Y%m%d-%H%M%S")
    business_slug = _safe_filename_part(getattr(business, "slug", "") or getattr(business, "name", "business"))

    workbook = Workbook()
    manifest = workbook.active
    manifest.title = "Manifest"
    manifest.append(["MoneyPro Business Backup"])
    manifest.append(["Business", str(business)])
    manifest.append(["Business ID", business.pk])
    manifest.append(["Exported At", now_dt.isoformat(sep=" ", timespec="seconds")])
    manifest.append([])
    manifest.append(["Sheet", "Model", "Database Table", "Rows"])

    used_names = {"Manifest"}
    header_fill = PatternFill("solid", fgColor="E9ECEF")
    header_font = Font(bold=True)

    for spec in specs:
        queryset = queryset_for_business(spec, business)
        headers, rows = rows_for_queryset(queryset)
        sheet_name = _unique_sheet_name(spec.model_name, used_names)
        sheet = workbook.create_sheet(title=sheet_name)

        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = header_font
            cell.fill = header_fill

        for row in rows:
            sheet.append(row)

        sheet.freeze_panes = "A2"
        for index, header in enumerate(headers, start=1):
            max_len = len(str(header))
            for row in rows[:100]:
                if index <= len(row):
                    max_len = max(max_len, len(str(row[index - 1])))
            sheet.column_dimensions[get_column_letter(index)].width = min(max(max_len + 2, 10), 45)

        manifest.append([sheet_name, spec.label, spec.table_name, len(rows)])

    for cell in manifest[6]:
        cell.font = header_font
        cell.fill = header_fill
    manifest.freeze_panes = "A7"
    manifest.column_dimensions["A"].width = 28
    manifest.column_dimensions["B"].width = 32
    manifest.column_dimensions["C"].width = 34
    manifest.column_dimensions["D"].width = 12

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)

    return buffer.getvalue()


def workbook_response_for_business(*, business) -> HttpResponse:
    now = timezone.localtime(timezone.now()).strftime("%Y%m%d-%H%M%S")
    business_slug = _safe_filename_part(getattr(business, "slug", "") or getattr(business, "name", "business"))
    data = workbook_bytes_for_business(business=business)
    response = HttpResponse(
        data,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{business_slug}-moneypro-backup-{now}.xlsx"'
    return response
